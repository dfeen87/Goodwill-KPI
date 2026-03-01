"""
Tests for the Goodwill Dashboard HTTP service.

Coverage:
- GET /health returns 200 with {"status": "ok"}.
- GET /dashboard returns 200 HTML response.
- POST /api/goodwill/calculate returns correct G, CG, UGS values and breakdowns.
- POST /api/goodwill/calculate with out-of-range inputs returns 422.
- POST /api/goodwill/calculate with T=0 returns 422.
- Response includes term contributions, penalty labels, and time-normalization flags.
- Results are fully traceable to inputs (deterministic).
- POST /api/goodwill/export returns xlsx file download by default.
- POST /api/goodwill/export?format=csv returns csv file download.
- Export file contains all required sections.
- Export rejects invalid inputs (422) and invalid format (422).
"""

import pytest
from fastapi.testclient import TestClient

from app.main import app

client = TestClient(app)

# Shared valid payload for reuse across tests
_VALID_PAYLOAD = {
    "CR": 80, "ES": 75, "BT": 70, "RG": 65, "NCB": 10,
    "CS": 80, "BR": 75, "CA": 70, "SS": 65, "NCB_consumer": 5,
    "T": 4,
    "g_w1": 0.25, "g_w2": 0.25, "g_w3": 0.25, "g_w_t": 0.25,
    "cg_w1": 0.25, "cg_w2": 0.25, "cg_w3": 0.25, "cg_w4": 0.25,
    "ugs_w1": 0.5, "ugs_w2": 0.5,
}


# ---------------------------------------------------------------------------
# /health
# ---------------------------------------------------------------------------


def test_health_returns_200():
    resp = client.get("/health")
    assert resp.status_code == 200
    assert resp.json() == {"status": "ok"}


# ---------------------------------------------------------------------------
# /dashboard
# ---------------------------------------------------------------------------


def test_dashboard_returns_html():
    resp = client.get("/dashboard")
    assert resp.status_code == 200
    assert "text/html" in resp.headers["content-type"]


def test_dashboard_contains_governance_warning():
    resp = client.get("/dashboard")
    assert "comparative indicators" in resp.text.lower()


# ---------------------------------------------------------------------------
# /api/goodwill/calculate — correct values
# ---------------------------------------------------------------------------


def test_calculate_known_values():
    # Manual verification:
    # G = ((80*0.25)+(75*0.25)+(70*0.25)+(65*0.25) - 10) / 4
    #   = (20+18.75+17.5+16.25 - 10) / 4 = 62.5 / 4 = 15.625
    # CG = (80*0.25)+(75*0.25)+(70*0.25)+(65*0.25) - 5
    #    = 20+18.75+17.5+16.25 - 5 = 67.5
    # UGS = ((15.625*0.5)+(67.5*0.5)) / 4 = 41.5625 / 4 = 10.390625
    resp = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD)
    assert resp.status_code == 200
    data = resp.json()
    assert data["G"] == pytest.approx(15.625)
    assert data["CG"] == pytest.approx(67.5)
    assert data["UGS"] == pytest.approx(10.390625)


def test_calculate_is_deterministic():
    r1 = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD).json()
    r2 = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD).json()
    assert r1["G"] == r2["G"]
    assert r1["CG"] == r2["CG"]
    assert r1["UGS"] == r2["UGS"]


# ---------------------------------------------------------------------------
# /api/goodwill/calculate — term breakdowns
# ---------------------------------------------------------------------------


def test_calculate_includes_G_term_breakdown():
    data = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD).json()
    terms = data["G_terms"]
    assert "CR_term" in terms
    assert "ES_term" in terms
    assert "BT_term" in terms
    assert "RG_term" in terms
    assert "NCB_penalty" in terms
    assert terms["time_normalized"] is True
    assert terms["T"] == _VALID_PAYLOAD["T"]


def test_calculate_includes_CG_term_breakdown():
    data = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD).json()
    terms = data["CG_terms"]
    assert "CS_term" in terms
    assert "BR_term" in terms
    assert "CA_term" in terms
    assert "SS_term" in terms
    assert "NCB_consumer_penalty" in terms
    assert terms["time_normalized"] is False


def test_calculate_includes_UGS_term_breakdown():
    data = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD).json()
    terms = data["UGS_terms"]
    assert "G_term" in terms
    assert "CG_term" in terms
    assert terms["time_normalized"] is True
    assert terms["T"] == _VALID_PAYLOAD["T"]


def test_calculate_NCB_penalty_reduces_G():
    low_ncb = {**_VALID_PAYLOAD, "NCB": 0}
    high_ncb = {**_VALID_PAYLOAD, "NCB": 50}
    g_clean = client.post("/api/goodwill/calculate", json=low_ncb).json()["G"]
    g_dirty = client.post("/api/goodwill/calculate", json=high_ncb).json()["G"]
    assert g_dirty < g_clean


def test_calculate_NCB_consumer_penalty_reduces_CG():
    low_ncb = {**_VALID_PAYLOAD, "NCB_consumer": 0}
    high_ncb = {**_VALID_PAYLOAD, "NCB_consumer": 50}
    cg_clean = client.post("/api/goodwill/calculate", json=low_ncb).json()["CG"]
    cg_dirty = client.post("/api/goodwill/calculate", json=high_ncb).json()["CG"]
    assert cg_dirty < cg_clean


def test_calculate_time_normalization_applied_once():
    # Doubling T should halve G (instantaneous CG is unaffected)
    # UGS is NOT simply halved because G itself changes with T (applied once in compute_G)
    t1 = {**_VALID_PAYLOAD, "T": 2}
    t2 = {**_VALID_PAYLOAD, "T": 4}
    r1 = client.post("/api/goodwill/calculate", json=t1).json()
    r2 = client.post("/api/goodwill/calculate", json=t2).json()
    assert r2["G"] == pytest.approx(r1["G"] / 2)
    assert r2["CG"] == pytest.approx(r1["CG"])  # CG is instantaneous — not T-normalized


def test_calculate_weights_used_present():
    data = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD).json()
    weights = data["weights_used"]
    for key in ("g_w1", "g_w2", "g_w3", "g_w_t", "cg_w1", "cg_w2", "cg_w3", "cg_w4",
                "ugs_w1", "ugs_w2"):
        assert key in weights


# ---------------------------------------------------------------------------
# /api/goodwill/calculate — input validation
# ---------------------------------------------------------------------------


def test_calculate_metric_above_100_returns_422():
    resp = client.post("/api/goodwill/calculate", json={**_VALID_PAYLOAD, "CR": 101})
    assert resp.status_code == 422


def test_calculate_metric_below_0_returns_422():
    resp = client.post("/api/goodwill/calculate", json={**_VALID_PAYLOAD, "ES": -1})
    assert resp.status_code == 422


def test_calculate_T_zero_returns_422():
    resp = client.post("/api/goodwill/calculate", json={**_VALID_PAYLOAD, "T": 0})
    assert resp.status_code == 422


def test_calculate_T_negative_returns_422():
    resp = client.post("/api/goodwill/calculate", json={**_VALID_PAYLOAD, "T": -1})
    assert resp.status_code == 422


def test_calculate_missing_required_field_returns_422():
    payload = {k: v for k, v in _VALID_PAYLOAD.items() if k != "CR"}
    resp = client.post("/api/goodwill/calculate", json=payload)
    assert resp.status_code == 422


# ---------------------------------------------------------------------------
# /api/goodwill/export — xlsx (default)
# ---------------------------------------------------------------------------


def test_export_xlsx_returns_200():
    resp = client.post("/api/goodwill/export", json=_VALID_PAYLOAD)
    assert resp.status_code == 200


def test_export_xlsx_content_type():
    resp = client.post("/api/goodwill/export", json=_VALID_PAYLOAD)
    assert "spreadsheetml" in resp.headers["content-type"]


def test_export_xlsx_filename_in_disposition():
    resp = client.post("/api/goodwill/export", json=_VALID_PAYLOAD)
    disposition = resp.headers.get("content-disposition", "")
    assert "goodwill_export_" in disposition
    assert ".xlsx" in disposition


def test_export_xlsx_is_valid_workbook():
    import io
    import openpyxl
    resp = client.post("/api/goodwill/export", json=_VALID_PAYLOAD)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Goodwill Export" in wb.sheetnames


def test_export_xlsx_contains_required_sections():
    import io
    import openpyxl
    resp = client.post("/api/goodwill/export", json=_VALID_PAYLOAD)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Goodwill Export"]
    sections = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert "G Inputs" in sections
    assert "CG Inputs" in sections
    assert "Weights & Time" in sections
    assert "Term Breakdown" in sections
    assert "Final Results" in sections


def test_export_xlsx_final_results_match_calculate():
    import io
    import openpyxl
    calc = client.post("/api/goodwill/calculate", json=_VALID_PAYLOAD).json()
    resp = client.post("/api/goodwill/export", json=_VALID_PAYLOAD)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Goodwill Export"]
    results = {
        ws.cell(row=r, column=2).value: ws.cell(row=r, column=3).value
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "Final Results"
    }
    assert results["G (General Goodwill)"] == pytest.approx(calc["G"])
    assert results["CG (Consumer Goodwill)"] == pytest.approx(calc["CG"])
    assert results["UGS (Unified Goodwill Score)"] == pytest.approx(calc["UGS"])


def test_export_xlsx_ncb_penalty_is_negative():
    import io
    import openpyxl
    resp = client.post("/api/goodwill/export", json=_VALID_PAYLOAD)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Goodwill Export"]
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value or ""
        if "NCB penalty" in label or "NCB_consumer penalty" in label:
            assert ws.cell(row=r, column=3).value <= 0


# ---------------------------------------------------------------------------
# /api/goodwill/export?format=csv
# ---------------------------------------------------------------------------


def test_export_csv_returns_200():
    resp = client.post("/api/goodwill/export?format=csv", json=_VALID_PAYLOAD)
    assert resp.status_code == 200


def test_export_csv_content_type():
    resp = client.post("/api/goodwill/export?format=csv", json=_VALID_PAYLOAD)
    assert "text/csv" in resp.headers["content-type"]


def test_export_csv_filename_in_disposition():
    resp = client.post("/api/goodwill/export?format=csv", json=_VALID_PAYLOAD)
    disposition = resp.headers.get("content-disposition", "")
    assert "goodwill_export_" in disposition
    assert ".csv" in disposition


def test_export_csv_contains_required_sections():
    import csv
    import io
    resp = client.post("/api/goodwill/export?format=csv", json=_VALID_PAYLOAD)
    reader = csv.DictReader(io.StringIO(resp.text))
    sections = {row["Section"] for row in reader}
    assert "G Inputs" in sections
    assert "CG Inputs" in sections
    assert "Weights & Time" in sections
    assert "Term Breakdown" in sections
    assert "Final Results" in sections


# ---------------------------------------------------------------------------
# /api/goodwill/export — input validation
# ---------------------------------------------------------------------------


def test_export_invalid_format_returns_422():
    resp = client.post("/api/goodwill/export?format=pdf", json=_VALID_PAYLOAD)
    assert resp.status_code == 422


def test_export_missing_required_field_returns_422():
    payload = {k: v for k, v in _VALID_PAYLOAD.items() if k != "CR"}
    resp = client.post("/api/goodwill/export", json=payload)
    assert resp.status_code == 422


def test_export_metric_above_100_returns_422():
    resp = client.post("/api/goodwill/export", json={**_VALID_PAYLOAD, "CR": 101})
    assert resp.status_code == 422


def test_export_T_zero_returns_422():
    resp = client.post("/api/goodwill/export", json={**_VALID_PAYLOAD, "T": 0})
    assert resp.status_code == 422
