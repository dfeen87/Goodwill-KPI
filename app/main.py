"""Goodwill Dashboard — FastAPI application.

Routes
------
GET  /health                    → 200 OK (liveness probe)
GET  /dashboard                 → Read-only calculation interface (HTML)
POST /api/goodwill/calculate    → Execute goodwill equations server-side;
                                  returns G, CG, UGS and term breakdowns

Constraints
-----------
- All goodwill equations are executed server-side via the pure functions in
  ``goodwill.metrics``.  This module does NOT duplicate any equation logic.
- No state is stored, mutated, or persisted.
- No forecasting, optimization, or action suggestions are performed.
"""

from __future__ import annotations

import csv
import io
import math
import os
from datetime import datetime, timezone
from typing import NamedTuple, Optional

from fastapi import FastAPI, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field, model_validator

from goodwill.metrics import compute_CG, compute_G, compute_UGS
from goodwill import config as _cfg
from goodwill import __version__

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------

app = FastAPI(
    title="Goodwill KPI Dashboard",
    description="Read-only governance view for Goodwill metric calculations.",
    version=__version__,
)

_TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")
templates = Jinja2Templates(directory=_TEMPLATES_DIR)


# ---------------------------------------------------------------------------
# Security headers middleware
# ---------------------------------------------------------------------------


@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    """Add standard security headers to every response."""
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Content-Security-Policy"] = (
        # 'unsafe-inline' is required for the dashboard's inline <script> block.
        # Moving scripts to an external file would allow a stricter policy.
        "default-src 'self'; script-src 'unsafe-inline'"
    )
    return response


# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------


class GoodwillInput(BaseModel):
    """Normalized metric inputs for all three goodwill equations."""

    # General Goodwill (G) inputs — range [0, 100]
    CR: float = Field(..., ge=0, le=100, description="Customer Retention [0–100]")
    ES: float = Field(..., ge=0, le=100, description="Employee Satisfaction [0–100]")
    BT: float = Field(..., ge=0, le=100, description="Brand Trust [0–100]")
    RG: float = Field(..., ge=0, le=100, description="Revenue Growth [0–100]")
    NCB: float = Field(..., ge=0, le=100, description="Net Customer Backlash [0–100]")

    # Consumer Goodwill (CG) inputs — range [0, 100]
    CS: float = Field(..., ge=0, le=100, description="Customer Satisfaction [0–100]")
    BR: float = Field(..., ge=0, le=100, description="Brand Reputation [0–100]")
    CA: float = Field(..., ge=0, le=100, description="Customer Advocacy [0–100]")
    SS: float = Field(..., ge=0, le=100, description="Service Speed [0–100]")
    NCB_consumer: float = Field(
        ..., ge=0, le=100, description="Negative Consumer Backlash [0–100]"
    )

    # Time normalization factor — must be > 0
    T: float = Field(..., gt=0, description="Time normalization factor (T > 0)")

    # Optional weight overrides (G)
    g_w1: Optional[float] = Field(None, description="Weight for CR (default: config)")
    g_w2: Optional[float] = Field(None, description="Weight for ES (default: config)")
    g_w3: Optional[float] = Field(None, description="Weight for BT (default: config)")
    g_w_t: Optional[float] = Field(None, description="Weight w(t) for RG (default: config)")

    # Optional weight overrides (CG)
    cg_w1: Optional[float] = Field(None, description="Weight for CS (default: config)")
    cg_w2: Optional[float] = Field(None, description="Weight for BR (default: config)")
    cg_w3: Optional[float] = Field(None, description="Weight for CA (default: config)")
    cg_w4: Optional[float] = Field(None, description="Weight for SS (default: config)")

    # Optional weight overrides (UGS)
    ugs_w1: Optional[float] = Field(None, description="Weight for G in UGS (default: config)")
    ugs_w2: Optional[float] = Field(None, description="Weight for CG in UGS (default: config)")

    @model_validator(mode="after")
    def _validate_weights_finite(self) -> "GoodwillInput":
        """Reject non-finite weight overrides (NaN, Infinity) at the API boundary."""
        weight_fields = (
            "g_w1", "g_w2", "g_w3", "g_w_t",
            "cg_w1", "cg_w2", "cg_w3", "cg_w4",
            "ugs_w1", "ugs_w2",
        )
        for field_name in weight_fields:
            value = getattr(self, field_name)
            if value is not None and not math.isfinite(value):
                raise ValueError(
                    f"Weight '{field_name}' = {value!r} must be a finite real number."
                )
        return self


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


@app.get("/", include_in_schema=False)
def root() -> RedirectResponse:
    """Redirect root URL to the dashboard."""
    return RedirectResponse(url="/dashboard", status_code=302)


@app.get("/health", status_code=200)
def health() -> dict:
    """Liveness probe — returns 200 OK."""
    return {"status": "ok"}


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request) -> HTMLResponse:
    """Render the read-only Goodwill calculation interface."""
    return templates.TemplateResponse(request, "dashboard.html")


@app.post("/api/goodwill/calculate")
def calculate(inputs: GoodwillInput) -> JSONResponse:
    """Execute all goodwill equations server-side and return full breakdowns.

    All calculations are delegated to the pure functions in ``goodwill.metrics``.
    No state is persisted.  Results are fully traceable to inputs.
    """
    # Resolve weights — use explicit override or fall back to config defaults
    weights = _resolve_weights(inputs)

    # Execute equations (server-side, via pure functions — no duplication)
    return JSONResponse(content=_build_result(inputs, weights))


# ---------------------------------------------------------------------------
# Shared calculation helper
# ---------------------------------------------------------------------------


class ResolvedWeights(NamedTuple):
    """Named container for resolved weight values used across all equations."""
    g_w1: float
    g_w2: float
    g_w3: float
    g_w_t: float
    cg_w1: float
    cg_w2: float
    cg_w3: float
    cg_w4: float
    ugs_w1: float
    ugs_w2: float


def _resolve_weights(inputs: GoodwillInput) -> ResolvedWeights:
    """Return resolved weight values from inputs or config defaults."""
    return ResolvedWeights(
        g_w1=inputs.g_w1 if inputs.g_w1 is not None else _cfg.G_W1,
        g_w2=inputs.g_w2 if inputs.g_w2 is not None else _cfg.G_W2,
        g_w3=inputs.g_w3 if inputs.g_w3 is not None else _cfg.G_W3,
        g_w_t=inputs.g_w_t if inputs.g_w_t is not None else _cfg.G_W_T,
        cg_w1=inputs.cg_w1 if inputs.cg_w1 is not None else _cfg.CG_W1,
        cg_w2=inputs.cg_w2 if inputs.cg_w2 is not None else _cfg.CG_W2,
        cg_w3=inputs.cg_w3 if inputs.cg_w3 is not None else _cfg.CG_W3,
        cg_w4=inputs.cg_w4 if inputs.cg_w4 is not None else _cfg.CG_W4,
        ugs_w1=inputs.ugs_w1 if inputs.ugs_w1 is not None else _cfg.UGS_W1,
        ugs_w2=inputs.ugs_w2 if inputs.ugs_w2 is not None else _cfg.UGS_W2,
    )


def _build_result(inputs: GoodwillInput, weights: ResolvedWeights) -> dict:
    """Execute all goodwill equations and return the full result payload."""
    G = compute_G(
        CR=inputs.CR, ES=inputs.ES, BT=inputs.BT, RG=inputs.RG,
        NCB=inputs.NCB, T=inputs.T,
        w1=weights.g_w1, w2=weights.g_w2, w3=weights.g_w3, w_t=weights.g_w_t,
    )
    CG = compute_CG(
        CS=inputs.CS, BR=inputs.BR, CA=inputs.CA, SS=inputs.SS,
        NCB_consumer=inputs.NCB_consumer,
        w1=weights.cg_w1, w2=weights.cg_w2, w3=weights.cg_w3, w4=weights.cg_w4,
    )
    UGS = compute_UGS(G=G, CG=CG, T=inputs.T, w1=weights.ugs_w1, w2=weights.ugs_w2)
    return {
        "G": G,
        "CG": CG,
        "UGS": UGS,
        "G_terms": {
            "CR_term": inputs.CR * weights.g_w1,
            "ES_term": inputs.ES * weights.g_w2,
            "BT_term": inputs.BT * weights.g_w3,
            "RG_term": inputs.RG * weights.g_w_t,
            "NCB_penalty": inputs.NCB,
            "T": inputs.T,
            "time_normalized": True,
        },
        "CG_terms": {
            "CS_term": inputs.CS * weights.cg_w1,
            "BR_term": inputs.BR * weights.cg_w2,
            "CA_term": inputs.CA * weights.cg_w3,
            "SS_term": inputs.SS * weights.cg_w4,
            "NCB_consumer_penalty": inputs.NCB_consumer,
            "time_normalized": False,
        },
        "UGS_terms": {
            "G_term": G * weights.ugs_w1,
            "CG_term": CG * weights.ugs_w2,
            "T": inputs.T,
            "time_normalized": True,
        },
        "weights_used": {
            "g_w1": weights.g_w1, "g_w2": weights.g_w2,
            "g_w3": weights.g_w3, "g_w_t": weights.g_w_t,
            "cg_w1": weights.cg_w1, "cg_w2": weights.cg_w2,
            "cg_w3": weights.cg_w3, "cg_w4": weights.cg_w4,
            "ugs_w1": weights.ugs_w1, "ugs_w2": weights.ugs_w2,
        },
    }


# ---------------------------------------------------------------------------
# Export endpoint
# ---------------------------------------------------------------------------


@app.post("/api/goodwill/export")
def export(
    inputs: GoodwillInput,
    format: str = Query(default="xlsx", pattern="^(xlsx|csv)$"),
) -> StreamingResponse:
    """Return a file download (xlsx or csv) of the full Goodwill calculation.

    The payload is identical to ``/api/goodwill/calculate``.  An optional
    ``format`` query parameter selects the output format (default: ``xlsx``).
    No additional computation beyond the standard equations is performed.
    """
    weights = _resolve_weights(inputs)
    d = _build_result(inputs, weights)

    # 'format' is regex-validated to 'xlsx|csv' by the Query constraint, so the
    # filename is safe to use directly in the Content-Disposition header.
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H%M%S")
    filename = f"goodwill_export_{timestamp}.{format}"

    if format == "csv":
        return _build_csv_response(inputs, d, filename)
    return _build_xlsx_response(inputs, d, filename)


# ---------------------------------------------------------------------------
# File-building helpers
# ---------------------------------------------------------------------------


def _export_rows(inputs: GoodwillInput, d: dict) -> list[tuple[str, str, float]]:
    """Return ordered (section, label, value) triples for the export file."""
    w = d["weights_used"]
    return [
        # --- Inputs: G ---
        ("G Inputs", "CR (Customer Retention)", inputs.CR),
        ("G Inputs", "ES (Employee Satisfaction)", inputs.ES),
        ("G Inputs", "BT (Brand Trust)", inputs.BT),
        ("G Inputs", "RG (Revenue Growth)", inputs.RG),
        ("G Inputs", "NCB (Net Customer Backlash)", inputs.NCB),
        # --- Inputs: CG ---
        ("CG Inputs", "CS (Customer Satisfaction)", inputs.CS),
        ("CG Inputs", "BR (Brand Reputation)", inputs.BR),
        ("CG Inputs", "CA (Customer Advocacy)", inputs.CA),
        ("CG Inputs", "SS (Service Speed)", inputs.SS),
        ("CG Inputs", "NCB_consumer (Negative Consumer Backlash)", inputs.NCB_consumer),
        # --- Weights & Time ---
        ("Weights & Time", "T (Time normalization factor)", inputs.T),
        ("Weights & Time", "g_w1 (weight for CR)", w["g_w1"]),
        ("Weights & Time", "g_w2 (weight for ES)", w["g_w2"]),
        ("Weights & Time", "g_w3 (weight for BT)", w["g_w3"]),
        ("Weights & Time", "g_w_t (weight for RG)", w["g_w_t"]),
        ("Weights & Time", "cg_w1 (weight for CS)", w["cg_w1"]),
        ("Weights & Time", "cg_w2 (weight for BR)", w["cg_w2"]),
        ("Weights & Time", "cg_w3 (weight for CA)", w["cg_w3"]),
        ("Weights & Time", "cg_w4 (weight for SS)", w["cg_w4"]),
        ("Weights & Time", "ugs_w1 (weight for G)", w["ugs_w1"]),
        ("Weights & Time", "ugs_w2 (weight for CG)", w["ugs_w2"]),
        # --- Term Breakdown ---
        ("Term Breakdown", "CR × w1", d["G_terms"]["CR_term"]),
        ("Term Breakdown", "ES × w2", d["G_terms"]["ES_term"]),
        ("Term Breakdown", "BT × w3", d["G_terms"]["BT_term"]),
        ("Term Breakdown", "RG × w(t)", d["G_terms"]["RG_term"]),
        ("Term Breakdown", "− NCB penalty", -d["G_terms"]["NCB_penalty"]),
        ("Term Breakdown", "CS × w1", d["CG_terms"]["CS_term"]),
        ("Term Breakdown", "BR × w2", d["CG_terms"]["BR_term"]),
        ("Term Breakdown", "CA × w3", d["CG_terms"]["CA_term"]),
        ("Term Breakdown", "SS × w4", d["CG_terms"]["SS_term"]),
        ("Term Breakdown", "− NCB_consumer penalty", -d["CG_terms"]["NCB_consumer_penalty"]),
        ("Term Breakdown", "G × ugs_w1", d["UGS_terms"]["G_term"]),
        ("Term Breakdown", "CG × ugs_w2", d["UGS_terms"]["CG_term"]),
        # --- Final Results ---
        ("Final Results", "G (General Goodwill)", d["G"]),
        ("Final Results", "CG (Consumer Goodwill)", d["CG"]),
        ("Final Results", "UGS (Unified Goodwill Score)", d["UGS"]),
    ]


def _build_xlsx_response(inputs: GoodwillInput, d: dict, filename: str) -> StreamingResponse:
    """Build an xlsx StreamingResponse from the result data."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Goodwill Export"

    # Header row
    ws.append(["Section", "Label", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A2D47")

    section_fill = {
        "G Inputs": "DBEAFE",
        "CG Inputs": "DCFCE7",
        "Weights & Time": "FEF9C3",
        "Term Breakdown": "F3E8FF",
        "Final Results": "FEE2E2",
    }

    for section, label, value in _export_rows(inputs, d):
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=section)
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=value)
        fill_color = section_fill.get(section, "FFFFFF")
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=fill_color)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_csv_response(inputs: GoodwillInput, d: dict, filename: str) -> StreamingResponse:
    """Build a CSV StreamingResponse from the result data."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Section", "Label", "Value"])
    writer.writerows(_export_rows(inputs, d))
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
